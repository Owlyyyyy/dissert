from __future__ import annotations

import argparse
import csv
import re
from pathlib import Path

from openpyxl import load_workbook


SECONDS_RE = re.compile(r"^\s*(\d+(?:\.\d+)?)\s*$")
MINSEC_RE = re.compile(r"^\s*(\d+)\s*:\s*(\d+(?:\.\d+)?)\s*$")


def norm(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_time(token: str) -> float:
    token = token.strip()
    match = SECONDS_RE.fullmatch(token)
    if match:
        return float(match.group(1))
    match = MINSEC_RE.fullmatch(token)
    if match:
        return int(match.group(1)) * 60 + float(match.group(2))
    raise ValueError(f"Invalid time token: {token!r}")


def normalize_intervals(value) -> str:
    text = norm(value)
    if not text:
        return ""
    parts = []
    for raw_part in text.split(";"):
        part = raw_part.strip()
        if not part:
            continue
        if "-" not in part:
            raise ValueError(f"Invalid interval: {part!r}")
        start_text, end_text = part.split("-", 1)
        start = parse_time(start_text)
        end = parse_time(end_text)
        if end <= start:
            raise ValueError(f"Interval end must be after start: {part!r}")
        parts.append(f"{start:.2f}-{end:.2f}")
    return ";".join(parts)


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


def write_csv_rows(path: Path, rows: list[dict[str, str]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def header_map(ws) -> dict[str, int]:
    headers = [norm(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    return {header: idx for idx, header in enumerate(headers)}


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", required=True, type=Path)
    parser.add_argument("--source-csv", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    args = parser.parse_args()

    source_rows = read_csv_rows(args.source_csv)
    by_id = {row["sample_id"]: dict(row) for row in source_rows}
    fieldnames = list(source_rows[0].keys())

    wb = load_workbook(args.workbook, data_only=False, read_only=True)
    ws = wb["Annotations"]
    col = header_map(ws)

    required = [
        "Sample ID",
        "Chorus present",
        "Chorus intervals (s)",
        "Pre-chorus intervals (s)",
        "Post-chorus intervals (s)",
        "Reviewer",
        "Notes",
    ]
    missing = [name for name in required if name not in col]
    if missing:
        raise SystemExit(f"Missing workbook columns: {', '.join(missing)}")

    updated = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        sample_id = norm(row[col["Sample ID"]])
        if not sample_id:
            continue
        if sample_id not in by_id:
            raise SystemExit(f"Unknown sample id in workbook: {sample_id}")
        out = by_id[sample_id]
        out["human_chorus_present"] = norm(row[col["Chorus present"]]).casefold()
        out["human_chorus_intervals"] = normalize_intervals(row[col["Chorus intervals (s)"]])
        out["human_pre_chorus_intervals"] = normalize_intervals(row[col["Pre-chorus intervals (s)"]])
        out["human_post_chorus_intervals"] = normalize_intervals(row[col["Post-chorus intervals (s)"]])
        out["reviewer"] = norm(row[col["Reviewer"]])
        out["notes"] = norm(row[col["Notes"]])
        updated += 1

    rows = [by_id[row["sample_id"]] for row in source_rows]
    write_csv_rows(args.output, rows, fieldnames)
    print(f"Wrote {args.output}")
    print(f"Updated annotation rows: {updated}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
