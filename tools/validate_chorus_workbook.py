from __future__ import annotations

import argparse
import re
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


VALID_PRESENT = {"yes", "no", "uncertain", ""}
SECONDS_RE = re.compile(r"^\s*(-?\d+(?:\.\d+)?)\s*$")
MINSEC_RE = re.compile(r"^\s*(-?\d+)\s*:\s*(\d+(?:\.\d+)?)\s*$")


def norm(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_time_token(token: str):
    token = token.strip()
    seconds_match = SECONDS_RE.match(token)
    if seconds_match:
        return float(seconds_match.group(1)), False
    minsec_match = MINSEC_RE.match(token)
    if minsec_match:
        minutes = int(minsec_match.group(1))
        seconds = float(minsec_match.group(2))
        return minutes * 60 + seconds, True
    raise ValueError(token)


def parse_intervals(text: str):
    text = norm(text)
    if not text:
        return [], []
    intervals = []
    errors = []
    for part in text.split(";"):
        part = part.strip()
        if not part:
            continue
        if "-" not in part:
            errors.append(f"bad interval syntax: {part!r}")
            continue
        start_text, end_text = part.split("-", 1)
        try:
            start, _start_was_minsec = parse_time_token(start_text)
            end, _end_was_minsec = parse_time_token(end_text)
        except ValueError:
            errors.append(f"bad interval syntax: {part!r}")
            continue
        if start < 0 or end < 0:
            errors.append(f"negative interval: {part!r}")
        if end <= start:
            errors.append(f"end <= start: {part!r}")
        # We allow >30 because a few providers can expose slightly longer previews,
        # but flag it because this project mainly uses ~30s preview clips.
        if end > 35:
            errors.append(f"interval beyond typical preview length: {part!r}")
        intervals.append((start, end))
    return intervals, errors


def find_col(headers, wanted: str):
    wanted_norm = wanted.casefold().strip()
    for idx, header in enumerate(headers, start=1):
        if norm(header).casefold().strip() == wanted_norm:
            return idx
    return None


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    args = parser.parse_args()

    path = args.workbook
    wb = load_workbook(path, data_only=False, read_only=True)
    print(f"Workbook: {path}")
    print(f"Sheets: {', '.join(wb.sheetnames)}")

    if "Annotations" not in wb.sheetnames:
        print("ERROR: Missing 'Annotations' sheet")
        return 2

    ws = wb["Annotations"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    cols = {
        "artist": find_col(headers, "Artist"),
        "title": find_col(headers, "Title"),
        "present": find_col(headers, "Chorus present"),
        "chorus": find_col(headers, "Chorus intervals (s)"),
        "pre": find_col(headers, "Pre-chorus intervals (s)"),
        "post": find_col(headers, "Post-chorus intervals (s)"),
        "notes": find_col(headers, "Notes"),
    }
    missing = [name for name, col in cols.items() if name in {"present", "chorus"} and col is None]
    if missing:
        print(f"ERROR: Missing required columns: {', '.join(missing)}")
        return 2

    total_rows = 0
    filled_rows = 0
    present_counts = Counter()
    issues = []
    no_with_intervals = []
    yes_without_intervals = []
    uncertain_rows = []
    interval_total = 0
    duration_total = 0.0

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        values = {name: norm(row[col - 1]) if col else "" for name, col in cols.items()}
        if not any(values.values()):
            continue
        total_rows += 1
        present = values["present"].casefold()
        chorus_text = values["chorus"]
        artist = values["artist"] or "(missing artist)"
        title = values["title"] or "(missing title)"
        label = f"row {row_idx}: {artist} — {title}"

        if present or chorus_text:
            filled_rows += 1
        present_counts[present or "(blank)"] += 1

        if present not in VALID_PRESENT:
            issues.append(f"{label}: invalid Chorus present value {values['present']!r}")

        intervals, interval_errors = parse_intervals(chorus_text)
        for error in interval_errors:
            issues.append(f"{label}: {error}")
        interval_total += len(intervals)
        duration_total += sum(end - start for start, end in intervals)

        if present == "no" and intervals:
            no_with_intervals.append(label)
        if present == "yes" and not intervals:
            yes_without_intervals.append(label)
        if present == "uncertain":
            uncertain_rows.append(label)

        # Optional diagnostic interval columns: validate syntax if filled.
        for diag_name in ("pre", "post"):
            _, diag_errors = parse_intervals(values[diag_name])
            for error in diag_errors:
                issues.append(f"{label}: {diag_name} {error}")

    print(f"Annotation rows found: {total_rows}")
    print(f"Rows with annotation content: {filled_rows}")
    print("Chorus present counts:")
    for key, count in sorted(present_counts.items()):
        print(f"  {key}: {count}")
    print(f"Chorus intervals entered: {interval_total}")
    print(f"Total annotated chorus seconds: {duration_total:.2f}")

    if yes_without_intervals:
        print("Rows marked yes but missing chorus intervals:")
        for item in yes_without_intervals[:20]:
            print(f"  {item}")
        if len(yes_without_intervals) > 20:
            print(f"  ... {len(yes_without_intervals) - 20} more")

    if no_with_intervals:
        print("Rows marked no but containing chorus intervals:")
        for item in no_with_intervals[:20]:
            print(f"  {item}")
        if len(no_with_intervals) > 20:
            print(f"  ... {len(no_with_intervals) - 20} more")

    if uncertain_rows:
        print("Rows marked uncertain:")
        for item in uncertain_rows[:20]:
            print(f"  {item}")
        if len(uncertain_rows) > 20:
            print(f"  ... {len(uncertain_rows) - 20} more")

    if issues:
        print("Validation issues:")
        for issue in issues[:50]:
            print(f"  {issue}")
        if len(issues) > 50:
            print(f"  ... {len(issues) - 50} more")
        return 1

    print("Validation issues: none")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
